import openpyxl
import pandas as pd

file_path = "Courses & Categories - Final.xlsx"

# Check sheet names
wb = openpyxl.load_workbook(file_path)
print("Sheet names:", wb.sheetnames)

# Get the first sheet
ws = wb.active
print(f"\nActive sheet: {ws.title}")
print(f"Dimensions: {ws.dimensions}")

# Print all header columns
print("\nColumn Headers:")
for col_idx, cell in enumerate(ws[1], 1):
    print(f"  Column {col_idx}: {cell.value}")

# Print first few data rows
print("\nFirst 5 data rows:")
for row_idx in range(2, 7):
    row_data = []
    for col in ws.iter_cols(min_row=row_idx, max_row=row_idx):
        for cell in col:
            row_data.append(cell.value)
    print(f"  Row {row_idx}: {row_data}")
